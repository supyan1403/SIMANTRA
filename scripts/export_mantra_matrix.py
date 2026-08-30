import sys
import os
import json
import openpyxl

def generate_mantra_matrix(template_path, output_path, json_data_path):
    wb = openpyxl.load_workbook(template_path)
    
    if not os.path.exists(json_data_path):
        wb.save(output_path)
        return

    with open(json_data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # data = {
    #   'mode': 'blank' or 'filled',
    #   'tahun': 2026,
    #   'mitras': [{'nama': ..., 'id_sobat': ..., 'nik': ..., 'alamat': ..., 'pekerjaan': ..., 'jk': ...}],
    #   'kegiatans': [{'nama': ..., 'kode_mak': ..., 'bidang': ..., 'harga': ..., 'satuan': ...}],
    #   'allocations': {
    #       '1': [{'mitra_nama': ..., 'kegiatan_nama': ..., 'nominal': ...}], # 1 = Januari
    #       ...
    #   }
    # }

    month_sheet_names = {
        1: 'JANUARI', 2: 'FEBRUARI', 3: 'MARET', 4: 'APRIL',
        5: 'MEI', 6: 'JUNI', 7: 'JULI', 8: 'AGUSTUS',
        9: 'SEPTEMBR', 10: 'OKTOBER', 11: 'NOPEMBER', 12: 'DESEMBER'
    }

    # If filled mode, inject allocations into monthly sheets
    if data.get('mode') == 'filled' and 'allocations' in data:
        allocations = data['allocations']
        
        for m_num, sheet_name in month_sheet_names.items():
            str_m_num = str(m_num)
            if str_m_num not in allocations or sheet_name not in wb.sheetnames:
                continue
            
            sheet = wb[sheet_name]
            month_allocs = allocations[str_m_num]
            if not month_allocs:
                continue

            # Build map of kegiatan name -> column index from row 3
            keg_col_map = {}
            for col in range(11, sheet.max_column + 1):
                val = sheet.cell(row=3, column=col).value
                if val:
                    clean_keg = str(val).strip().lower()
                    keg_col_map[clean_keg] = col

            # Build map of mitra name -> row index
            mitra_sheet = wb['DB Mitra 2024 (2621)'] if 'DB Mitra 2024 (2621)' in wb.sheetnames else None
            mitra_row_map = {}
            
            if mitra_sheet:
                for r in range(3, min(mitra_sheet.max_row + 1, 2630)):
                    m_val = mitra_sheet.cell(row=r, column=2).value
                    if m_val:
                        clean_m = str(m_val).strip().lower()
                        # in month sheet, row index corresponds to r + 4 (row 3 in DB Mitra = row 7 in month sheet)
                        mitra_row_map[clean_m] = r + 4

            # Inject values
            for item in month_allocs:
                m_name = str(item.get('mitra_nama', '')).strip().lower()
                k_name = str(item.get('kegiatan_nama', '')).strip().lower()
                nominal = float(item.get('nominal', 0))

                target_row = mitra_row_map.get(m_name)
                target_col = None

                # Find best matching kegiatan column
                if k_name in keg_col_map:
                    target_col = keg_col_map[k_name]
                else:
                    for k_key, col_idx in keg_col_map.items():
                        if k_name in k_key or k_key in k_name:
                            target_col = col_idx
                            break

                if target_row and target_col:
                    sheet.cell(row=target_row, column=target_col, value=nominal)

    wb.save(output_path)

if __name__ == '__main__':
    if len(sys.argv) >= 4:
        template = sys.argv[1]
        output = sys.argv[2]
        json_path = sys.argv[3]
        generate_mantra_matrix(template, output, json_path)
    elif len(sys.argv) >= 3:
        template = sys.argv[1]
        output = sys.argv[2]
        generate_mantra_matrix(template, output, '')
