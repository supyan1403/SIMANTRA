import os
import sys
import sqlite3
import datetime
import openpyxl

db_path = 'd:/SIMANTRA/database/database.sqlite'
excel_path = 'd:/SIMANTRA/1. Input MANTRA (Monitoring Alokasi Pekerjaan Dan Honor Mitra) oleh PJ Kegiatan atau Ketua Tim - Update3.xlsx.xlsx'

if not os.path.exists(db_path):
    print(f"Error: Database not found at {db_path}")
    sys.exit(1)

if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    sys.exit(1)

conn = sqlite3.connect(db_path)
cur = conn.cursor()

now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

print("=== STEP 1: SETUP BIDANG & PERIODE 2024 ===")
bidang_names = ['Distribusi', 'Neraca', 'Produksi', 'Sosial', 'IPDS', 'Bagian Umum', 'Cadangan']
bidang_ids = {}
for b_name in bidang_names:
    cur.execute("SELECT id FROM bidangs WHERE LOWER(nama) = LOWER(?)", (b_name,))
    row = cur.fetchone()
    if row:
        bidang_ids[b_name.lower()] = row[0]
    else:
        cur.execute("INSERT INTO bidangs (nama, created_at, updated_at) VALUES (?, ?, ?)", (b_name, now_str, now_str))
        bidang_ids[b_name.lower()] = cur.lastrowid

default_bidang_id = bidang_ids.get('distribusi', 1)

def get_bidang_id(bidang_str):
    if not bidang_str:
        return default_bidang_id
    bs = str(bidang_str).lower().strip()
    for k, b_id in bidang_ids.items():
        if k in bs:
            return b_id
    return default_bidang_id

month_names = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
    5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
}
month_map = {
    'JANUARI': 1, 'FEBRUARI': 2, 'MARET': 3, 'APRIL': 4,
    'MEI': 5, 'JUNI': 6, 'JULI': 7, 'AGUSTUS': 8,
    'SEPTEMBR': 9, 'SEPTEMBER': 9, 'OKTOBER': 10, 'NOPEMBER': 11, 'NOVEMBER': 11, 'DESEMBER': 12
}

periode_ids = {}
for m_num, m_name in month_names.items():
    cur.execute("SELECT id FROM periodes WHERE tahun = '2024' AND (bulan = ? OR bulan_angka = ?)", (m_name, m_num))
    row = cur.fetchone()
    if row:
        periode_ids[m_num] = row[0]
    else:
        cur.execute("INSERT INTO periodes (tahun, bulan, bulan_angka, created_at, updated_at) VALUES ('2024', ?, ?, ?, ?)", (m_name, m_num, now_str, now_str))
        periode_ids[m_num] = cur.lastrowid

print("Loading Master Workbook...")
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

print("\n=== STEP 2: PEMBERSIHAN DATA KEGIATAN TIDAK VALID ===")
cur.execute("""
    SELECT id, nama FROM kegiatans
    WHERE tahun = '2023' OR total = 0 OR total IS NULL OR kode_mata_anggaran IS NULL OR kode_mata_anggaran = ''
""")
deleted_count = 0
for r in cur.fetchall():
    k_id, k_nama = r
    cur.execute("DELETE FROM alokasi_honors WHERE kegiatan_id = ?", (k_id,))
    cur.execute("DELETE FROM kegiatans WHERE id = ?", (k_id,))
    deleted_count += 1

print(f"Berhasil membersihkan {deleted_count} record kegiatan tidak valid.")

print("\n=== STEP 3: SINKRONISASI 89 MASTER KEGIATAN RESMI DARI 'DB MATA ANGGARAN' ===")
s_mak = wb['DB MATA ANGGARAN']
master_by_no = {}
master_by_name = {}
master_by_id = {}

for r_idx, row in enumerate(s_mak.iter_rows(values_only=True)):
    if r_idx == 0 or len(row) < 8 or not row[2] or str(row[2]).strip().upper() == 'KEGIATAN':
        continue
    
    no = row[0]
    bidang_str = str(row[1]).strip() if row[1] else 'Distribusi'
    b_id = get_bidang_id(bidang_str)
    
    nama = str(row[2]).strip()
    mak = str(row[3]).strip() if row[3] and not str(row[3]).startswith('=') else ''
    
    try:
        vol = int(round(float(row[4]))) if row[4] is not None else 0
    except:
        vol = 0
        
    satuan = str(row[5]).strip() if row[5] else 'Dokumen'
    
    try:
        tarif = float(row[6]) if row[6] is not None else 0
    except:
        tarif = 0
        
    try:
        pagu = float(row[7]) if row[7] is not None else (vol * tarif)
    except:
        pagu = vol * tarif
        
    cur.execute("SELECT id FROM kegiatans WHERE UPPER(TRIM(nama)) = ? AND (tahun = '2024' OR tahun IS NULL)", (nama.upper(),))
    existing_keg = cur.fetchone()
    
    if existing_keg:
        k_id = existing_keg[0]
        cur.execute("""
            UPDATE kegiatans
            SET bidang_id = ?, kode_mata_anggaran = ?, jumlah = ?, satuan = ?, harga = ?, total = ?, tahun = '2024', updated_at = ?
            WHERE id = ?
        """, (b_id, mak, vol, satuan, tarif, pagu, now_str, k_id))
    else:
        cur.execute("""
            INSERT INTO kegiatans (nama, bidang_id, kode_mata_anggaran, jumlah, satuan, harga, total, tahun, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, '2024', ?, ?)
        """, (nama, b_id, mak, vol, satuan, tarif, pagu, now_str, now_str))
        k_id = cur.lastrowid
        
    k_data = {
        'id': k_id,
        'no': no,
        'bidang_id': b_id,
        'nama': nama,
        'mak': mak,
        'vol': vol,
        'satuan': satuan,
        'tarif': tarif,
        'pagu': pagu
    }
    
    master_by_name[nama.upper()] = k_data
    master_by_id[k_id] = k_data
    if no is not None:
        try:
            master_by_no[int(float(no))] = k_data
        except:
            pass

print(f"Berhasil menyinkronkan {len(master_by_name)} Master Kegiatan Resmi!")

print("\n=== STEP 4: SINKRONISASI 2.621 MASTER MITRA DARI 'DB Mitra 2024 (2621)' (PEKERJAAN ASLI DARI KOLOM D) ===")
s_mitra = wb['DB Mitra 2024 (2621)']
mitra_by_name = {}
mitra_by_sobat = {}

cur.execute("SELECT id, UPPER(TRIM(nama)), id_sobat FROM mitras")
for r in cur.fetchall():
    if r[1]: mitra_by_name[r[1]] = r[0]
    if r[2]: mitra_by_sobat[str(r[2]).strip()] = r[0]

mitra_synced = 0
for r_idx, row in enumerate(s_mitra.iter_rows(values_only=True)):
    if r_idx < 2 or len(row) < 2:
        continue
    nama_raw = row[1] # Col B
    if not nama_raw or not str(nama_raw).strip() or str(nama_raw).strip().upper() in ['NAMA', 'NAMA LENGKAP']:
        continue
        
    nama = str(nama_raw).strip()
    alamat = str(row[2]).strip() if len(row) > 2 and row[2] and not str(row[2]).startswith('=') else None
    
    # REAL PEKERJAAN from Column D (index 3)
    pekerjaan_raw = str(row[3]).strip() if len(row) > 3 and row[3] and not str(row[3]).startswith('=') and not str(row[3]).startswith('#') else 'Lainnya/ Belum Bekerja'
    pekerjaan = pekerjaan_raw if pekerjaan_raw and pekerjaan_raw != '#N/A' else 'Lainnya/ Belum Bekerja'
    
    nik = str(row[6]).strip() if len(row) > 6 and row[6] and not str(row[6]).startswith('=') else None
    posisi = str(row[7]).strip() if len(row) > 7 and row[7] and not str(row[7]).startswith('=') else None
    email = str(row[9]).strip() if len(row) > 9 and row[9] and not str(row[9]).startswith('=') else None
    npwp = str(row[17]).strip() if len(row) > 17 and row[17] and not str(row[17]).startswith('=') else None
    jk_raw = str(row[18]).strip() if len(row) > 18 and row[18] else None
    jk = 'L' if jk_raw in ['1', 'L', 'l'] else ('P' if jk_raw in ['2', 'P', 'p'] else 'L')
    pendidikan = str(row[21]).strip() if len(row) > 21 and row[21] else None
    sobat_id = str(row[35]).strip() if len(row) > 35 and row[35] else None
    
    m_id = None
    if sobat_id and sobat_id in mitra_by_sobat:
        m_id = mitra_by_sobat[sobat_id]
    elif nama.upper() in mitra_by_name:
        m_id = mitra_by_name[nama.upper()]
        
    if m_id:
        cur.execute("""
            UPDATE mitras
            SET nama = ?, alamat = ?, pekerjaan = ?, nik = COALESCE(?, nik), posisi = ?,
                email = ?, npwp = ?, jk = ?, pendidikan = ?, id_sobat = COALESCE(?, id_sobat), updated_at = ?
            WHERE id = ?
        """, (nama, alamat, pekerjaan, nik, posisi, email, npwp, jk, pendidikan, sobat_id, now_str, m_id))
    else:
        cur.execute("""
            INSERT INTO mitras (nama, alamat, pekerjaan, nik, posisi, email, npwp, jk, pendidikan, id_sobat, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nama, alamat, pekerjaan, nik, posisi, email, npwp, jk, pendidikan, sobat_id, now_str, now_str))
        m_id = cur.lastrowid
        mitra_by_name[nama.upper()] = m_id
        if sobat_id:
            mitra_by_sobat[sobat_id] = m_id
            
    mitra_synced += 1

print(f"Berhasil menyinkronkan {mitra_synced} Data Master Mitra dengan Pekerjaan Asli!")

print("\n=== STEP 5: PROSES MATRIKS 12 BULAN (JANUARI - DESEMBER 2024) ===")
# Clear 2024 allocations
all_2024_p_ids = list(periode_ids.values())
placeholders = ','.join('?' * len(all_2024_p_ids))
cur.execute(f"DELETE FROM alokasi_honors WHERE periode_id IN ({placeholders})", all_2024_p_ids)

month_sheets = ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS', 'SEPTEMBR', 'OKTOBER', 'NOPEMBER', 'DESEMBER']

total_alokasi_inserted = 0
month_summary = {}

for ms in month_sheets:
    if ms not in wb.sheetnames:
        continue
    m_num = month_map.get(ms)
    p_id = periode_ids[m_num]
    
    sheet = wb[ms]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 7:
        continue
        
    col_kegiatan_map = {}
    for c_idx in range(10, min(len(rows[0]), 100)):
        keg_found = None
        col_names = [rows[r][c_idx] for r in range(min(6, len(rows))) if c_idx < len(rows[r])]
        
        # 1. Match by string name
        for val in col_names:
            if val and isinstance(val, str) and len(val.strip()) > 3:
                clean_v = val.strip().upper()
                if clean_v in master_by_name:
                    keg_found = master_by_name[clean_v]
                    break
                elif not clean_v.startswith('SBML') and clean_v not in ['DISTRIBUSI', 'PRODUKSI', 'NERACA', 'SOSIAL', 'IPDS']:
                    for m_name, m_data in master_by_name.items():
                        if clean_v in m_name or m_name in clean_v:
                            keg_found = m_data
                            break
                    if keg_found: break
                    
        # 2. Match by No Urut in row index 4 (Row 5)
        if not keg_found and len(rows) > 4 and c_idx < len(rows[4]):
            no_val = rows[4][c_idx]
            if no_val is not None:
                try:
                    no_int = int(float(no_val))
                    if no_int in master_by_no:
                        keg_found = master_by_no[no_int]
                except:
                    pass
                    
        if keg_found:
            col_kegiatan_map[c_idx] = keg_found
            
    # Process rows and aggregate per (mitra_id, kegiatan_id, p_id)
    monthly_alokasi = {}
    for row_idx in range(6, len(rows)):
        row = rows[row_idx]
        if len(row) < 2:
            continue
        nama_raw = row[1]
        if not nama_raw or not str(nama_raw).strip() or str(nama_raw).strip().upper() in ['NAMA', 'NAMA LENGKAP']:
            continue
            
        clean_nama = str(nama_raw).strip().upper()
        m_id = mitra_by_name.get(clean_nama)
        if not m_id:
            continue
            
        for c_idx, k_info in col_kegiatan_map.items():
            if c_idx < len(row):
                val = row[c_idx]
                if val is not None:
                    try:
                        num_val = float(val)
                    except:
                        num_val = 0
                        
                    if num_val > 0:
                        tarif = k_info['tarif']
                        satuan = k_info['satuan'] or 'Dokumen'
                        nominal = num_val # EXACT RUPIAH FROM EXCEL
                        
                        if tarif > 0:
                            volume = max(1, int(round(num_val / tarif)))
                        else:
                            volume = 1
                            
                        key = (m_id, k_info['id'], p_id)
                        if key in monthly_alokasi:
                            monthly_alokasi[key]['volume'] += volume
                            monthly_alokasi[key]['nominal'] += nominal
                        else:
                            monthly_alokasi[key] = {
                                'volume': volume,
                                'nominal': nominal,
                                'satuan': satuan,
                                'tarif': tarif
                            }

    # Insert aggregated monthly alokasi
    month_inserted = 0
    for (m_id, k_id, p_id), a_data in monthly_alokasi.items():
        cur.execute("""
            INSERT INTO alokasi_honors (mitra_id, kegiatan_id, periode_id, volume, nominal, satuan, tarif_satuan, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (m_id, k_id, p_id, a_data['volume'], a_data['nominal'], a_data['satuan'], a_data['tarif'], now_str, now_str))
        month_inserted += 1
        total_alokasi_inserted += 1
                        
    month_summary[ms] = {
        'mapped_cols': len(col_kegiatan_map),
        'inserted': month_inserted
    }

conn.commit()
conn.close()

print("\n=======================================================")
print("SINKRONISASI MURNI DATA MASTER SELESAI SUKSES!")
print(f"Total Alokasi Penugasan Tersimpan: {total_alokasi_inserted}")
print("=======================================================")
for ms, stat in month_summary.items():
    print(f" - {ms:10s} : {stat['inserted']:4d} Penugasan across {stat['mapped_cols']:2d} Kegiatan Resmi")
