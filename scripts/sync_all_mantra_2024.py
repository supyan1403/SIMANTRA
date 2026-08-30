import os
import sys
import sqlite3
import datetime
import openpyxl

db_path = 'd:/SIMANTRA/database/database.sqlite'
excel_path = 'd:/SIMANTRA/1. Input MANTRA (Monitoring Alokasi Pekerjaan Dan Honor Mitra) oleh PJ Kegiatan atau Ketua Tim - Update3.xlsx.xlsx'

if not os.path.exists(db_path):
    print(f"Error: Database not found at {db_path}")
    sys.exit(1)

if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    sys.exit(1)

conn = sqlite3.connect(db_path)
cur = conn.cursor()

now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# 1. Setup Bidang Map
bidang_names = ['Distribusi', 'Neraca', 'Produksi', 'Sosial', 'IPDS', 'Bagian Umum', 'Cadangan']
bidang_ids = {}
for b_name in bidang_names:
    cur.execute("SELECT id FROM bidangs WHERE LOWER(nama) = LOWER(?)", (b_name,))
    row = cur.fetchone()
    if row:
        bidang_ids[b_name.lower()] = row[0]
    else:
        cur.execute("INSERT INTO bidangs (nama, created_at, updated_at) VALUES (?, ?, ?)", (b_name, now_str, now_str))
        bidang_ids[b_name.lower()] = cur.lastrowid

default_bidang_id = bidang_ids.get('distribusi', 1)

def get_bidang_id(bidang_str):
    if not bidang_str:
        return default_bidang_id
    bs = str(bidang_str).lower()
    for k, b_id in bidang_ids.items():
        if k in bs:
            return b_id
    return default_bidang_id

# 2. Setup Periode Map for 2024
month_names = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
    5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
}
month_map = {
    'JANUARI': 1, 'FEBRUARI': 2, 'MARET': 3, 'APRIL': 4,
    'MEI': 5, 'JUNI': 6, 'JULI': 7, 'AGUSTUS': 8,
    'SEPTEMBR': 9, 'SEPTEMBER': 9, 'OKTOBER': 10, 'NOPEMBER': 11, 'NOVEMBER': 11, 'DESEMBER': 12
}

periode_ids = {}
for m_num, m_name in month_names.items():
    cur.execute("SELECT id FROM periodes WHERE tahun = '2024' AND (bulan = ? OR bulan_angka = ?)", (m_name, m_num))
    row = cur.fetchone()
    if row:
        periode_ids[m_num] = row[0]
    else:
        cur.execute("INSERT INTO periodes (tahun, bulan, bulan_angka, created_at, updated_at) VALUES ('2024', ?, ?, ?, ?)", (m_name, m_num, now_str, now_str))
        periode_ids[m_num] = cur.lastrowid

# 3. Load All Mitras into Cache
cur.execute("SELECT id, UPPER(TRIM(nama)), id_sobat, nik FROM mitras")
mitra_by_name = {}
mitra_by_sobat = {}
for r in cur.fetchall():
    m_id, m_name, m_sobat, m_nik = r
    if m_name:
        mitra_by_name[m_name] = m_id
    if m_sobat:
        mitra_by_sobat[str(m_sobat).strip()] = m_id

def get_or_create_mitra(nama_mitra, alamat=None, pekerjaan=None, jk=None):
    clean_name = str(nama_mitra).strip()
    upper_name = clean_name.upper()
    if upper_name in mitra_by_name:
        return mitra_by_name[upper_name]
    
    clean_pekerjaan = pekerjaan if pekerjaan and str(pekerjaan).strip() else 'Mitra Statistik'
    clean_jk = jk if jk in ['L', 'P'] else 'L'
    clean_alamat = alamat if alamat and str(alamat).strip() else None
    
    # Create new mitra
    cur.execute("""
        INSERT INTO mitras (nama, alamat, pekerjaan, jk, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (clean_name, clean_alamat, clean_pekerjaan, clean_jk, now_str, now_str))
    new_id = cur.lastrowid
    mitra_by_name[upper_name] = new_id
    return new_id

# 4. Load All Kegiatans into Cache
cur.execute("SELECT id, UPPER(TRIM(nama)), harga, kode_mata_anggaran, tahun, bidang_id FROM kegiatans WHERE tahun = '2024' OR tahun IS NULL")
kegiatan_by_name = {}
for r in cur.fetchall():
    k_id, k_name, k_harga, k_mak, k_tahun, k_bidang = r
    if k_name:
        kegiatan_by_name[k_name] = {
            'id': k_id,
            'harga': float(k_harga or 0),
            'mak': k_mak or '',
            'bidang_id': k_bidang
        }

def get_or_create_kegiatan(nama_kegiatan, tarif=0, mak=None, bidang_str=None):
    clean_name = str(nama_kegiatan).strip()
    upper_name = clean_name.upper()
    
    if upper_name in kegiatan_by_name:
        k_info = kegiatan_by_name[upper_name]
        # Update tarif/mak if missing
        if tarif > 0 and k_info['harga'] == 0:
            cur.execute("UPDATE kegiatans SET harga = ?, updated_at = ? WHERE id = ?", (tarif, now_str, k_info['id']))
            k_info['harga'] = tarif
        if mak and not k_info['mak']:
            cur.execute("UPDATE kegiatans SET kode_mata_anggaran = ?, updated_at = ? WHERE id = ?", (mak, now_str, k_info['id']))
            k_info['mak'] = mak
        return k_info['id']
    
    b_id = get_bidang_id(bidang_str)
    cur.execute("""
        INSERT INTO kegiatans (nama, bidang_id, kode_mata_anggaran, harga, satuan, tahun, created_at, updated_at)
        VALUES (?, ?, ?, ?, 'Dokumen', '2024', ?, ?)
    """, (clean_name, b_id, mak, tarif, now_str, now_str))
    new_id = cur.lastrowid
    kegiatan_by_name[upper_name] = {
        'id': new_id,
        'harga': float(tarif or 0),
        'mak': mak or '',
        'bidang_id': b_id
    }
    return new_id

print("Loading Master MANTRA Excel workbook...")
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

total_inserted = 0
total_updated = 0
month_summary = {}

for sname in wb.sheetnames:
    upper_sname = sname.upper().strip()
    m_num = month_map.get(upper_sname)
    if not m_num:
        continue
    
    sheet = wb[sname]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 7:
        continue
    
    p_id = periode_ids[m_num]
    r1 = rows[0] if len(rows) > 0 else [] # Bidang names
    r3 = rows[2] if len(rows) > 2 else [] # Kegiatan names
    r4 = rows[3] if len(rows) > 3 else [] # Tarif
    r5 = rows[4] if len(rows) > 4 else [] # MAK
    
    kegiatans_in_sheet = {}
    current_bidang = 'Distribusi'
    
    for col_idx in range(10, min(len(r3), 100)):
        b_cell = r1[col_idx] if col_idx < len(r1) else None
        if b_cell and str(b_cell).strip() and not str(b_cell).strip().upper() in ['PILIH TIM', 'STATUS SBML', 'JUMLAH TOTAL PENDAPATAN MITRA']:
            current_bidang = str(b_cell).strip()
            
        keg_name = r3[col_idx]
        if keg_name and str(keg_name).strip() and not str(keg_name).startswith('=') and '#REF!' not in str(keg_name):
            clean_keg_name = str(keg_name).strip()
            
            tarif_val = r4[col_idx] if col_idx < len(r4) else 0
            try:
                tarif = float(tarif_val) if tarif_val is not None else 0
            except:
                tarif = 0
                
            mak_val = r5[col_idx] if col_idx < len(r5) else ''
            mak = str(mak_val).strip() if mak_val and not str(mak_val).startswith('=') and '#REF!' not in str(mak_val) else None
            
            k_id = get_or_create_kegiatan(clean_keg_name, tarif=tarif, mak=mak, bidang_str=current_bidang)
            kegiatans_in_sheet[col_idx] = {
                'id': k_id,
                'nama': clean_keg_name,
                'tarif': tarif,
                'mak': mak
            }
    
    inserted_in_month = 0
    updated_in_month = 0
    
    for row_idx in range(6, len(rows)):
        row = rows[row_idx]
        if len(row) < 2:
            continue
        mitra_name = row[1]
        if not mitra_name or not str(mitra_name).strip() or str(mitra_name).strip().upper() in ['NAMA', 'NAMA LENGKAP']:
            continue
        
        clean_mitra = str(mitra_name).strip()
        alamat = str(row[2]).strip() if len(row) > 2 and row[2] and not str(row[2]).startswith('=') else None
        pekerjaan = str(row[3]).strip() if len(row) > 3 and row[3] and not str(row[3]).startswith('=') else None
        jk_raw = str(row[5]).strip() if len(row) > 5 and row[5] else None
        jk = 'L' if jk_raw in ['1', 'L', 'l'] else ('P' if jk_raw in ['2', 'P', 'p'] else None)
        
        m_id = get_or_create_mitra(clean_mitra, alamat=alamat, pekerjaan=pekerjaan, jk=jk)
        
        for col_idx, k_info in kegiatans_in_sheet.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None:
                    try:
                        num_val = float(val)
                    except:
                        num_val = 0
                    
                    if num_val > 0:
                        tarif = k_info['tarif']
                        if tarif > 0:
                            if num_val >= tarif:
                                volume = round(num_val / tarif, 2)
                                nominal = num_val
                            else:
                                volume = num_val
                                nominal = round(num_val * tarif, 2)
                        else:
                            volume = 1
                            nominal = num_val
                        
                        # Check existing alokasi
                        cur.execute("""
                            SELECT id, volume, nominal FROM alokasi_honors
                            WHERE mitra_id = ? AND kegiatan_id = ? AND periode_id = ?
                        """, (m_id, k_info['id'], p_id))
                        existing_alokasi = cur.fetchone()
                        
                        if existing_alokasi:
                            cur.execute("""
                                UPDATE alokasi_honors
                                SET volume = ?, nominal = ?, satuan = 'Dokumen', tarif_satuan = ?, updated_at = ?
                                WHERE id = ?
                            """, (volume, nominal, tarif, now_str, existing_alokasi[0]))
                            updated_in_month += 1
                            total_updated += 1
                        else:
                            cur.execute("""
                                INSERT INTO alokasi_honors (mitra_id, kegiatan_id, periode_id, volume, nominal, satuan, tarif_satuan, created_at, updated_at)
                                VALUES (?, ?, ?, ?, ?, 'Dokumen', ?, ?, ?)
                            """, (m_id, k_info['id'], p_id, volume, nominal, tarif, now_str, now_str))
                            inserted_in_month += 1
                            total_inserted += 1
                            
    month_summary[upper_sname] = {
        'inserted': inserted_in_month,
        'updated': updated_in_month,
        'total': inserted_in_month + updated_in_month
    }

# 5. Apply Specific SPK Details for LINA KARLINA
cur.execute("SELECT id FROM mitras WHERE UPPER(TRIM(nama)) = 'LINA KARLINA'")
lina_row = cur.fetchone()
if lina_row:
    lina_id = lina_row[0]
    p_maret = periode_ids[3] # Maret 2024
    
    # 1. SHKK (ID 1)
    k_shkk = kegiatan_by_name.get('PENDATAAN LAPANGAN PAKET KOMODITAS SHKK (NON PNS)')
    if k_shkk:
        shkk_id = k_shkk['id']
        cur.execute("""
            SELECT id FROM alokasi_honors WHERE mitra_id = ? AND kegiatan_id = ? AND periode_id = ?
        """, (lina_id, shkk_id, p_maret))
        shkk_alokasi = cur.fetchone()
        if shkk_alokasi:
            cur.execute("""
                UPDATE alokasi_honors
                SET volume = 15, nominal = 900000, tarif_satuan = 60000, satuan = 'Dokumen',
                    nomor_spk = '1001/PPK/SPK/03/2024', tanggal_spk = '2024-02-28', updated_at = ?
                WHERE id = ?
            """, (now_str, shkk_alokasi[0]))
        else:
            cur.execute("""
                INSERT INTO alokasi_honors (mitra_id, kegiatan_id, periode_id, volume, nominal, satuan, tarif_satuan, nomor_spk, tanggal_spk, created_at, updated_at)
                VALUES (?, ?, ?, 15, 900000, 'Dokumen', 60000, '1001/PPK/SPK/03/2024', '2024-02-28', ?, ?)
            """, (lina_id, shkk_id, p_maret, now_str, now_str))
            
    # 2. HKD (ID 4)
    k_hkd = kegiatan_by_name.get('PENDATAAN LAPANGAN SURVEI HARGA KONSUMEN PERDESAAN (HKD) NON PNS')
    if k_hkd:
        hkd_id = k_hkd['id']
        cur.execute("""
            SELECT id FROM alokasi_honors WHERE mitra_id = ? AND kegiatan_id = ? AND periode_id = ?
        """, (lina_id, hkd_id, p_maret))
        hkd_alokasi = cur.fetchone()
        if hkd_alokasi:
            cur.execute("""
                UPDATE alokasi_honors
                SET volume = 4, nominal = 260000, tarif_satuan = 65000, satuan = 'Dokumen',
                    nomor_spk = '1001/PPK/SPK/03/2024', tanggal_spk = '2024-02-28', updated_at = ?
                WHERE id = ?
            """, (now_str, hkd_alokasi[0]))
        else:
            cur.execute("""
                INSERT INTO alokasi_honors (mitra_id, kegiatan_id, periode_id, volume, nominal, satuan, tarif_satuan, nomor_spk, tanggal_spk, created_at, updated_at)
                VALUES (?, ?, ?, 4, 260000, 'Dokumen', 65000, '1001/PPK/SPK/03/2024', '2024-02-28', ?, ?)
            """, (lina_id, hkd_id, p_maret, now_str, now_str))

conn.commit()
conn.close()

print("\n=======================================================")
print(f"SINKRONISASI SELESAI SUKSES!")
print(f"Total Alokasi Baru Diinput : {total_inserted}")
print(f"Total Alokasi Diperbarui   : {total_updated}")
print(f"Total Keseluruhan Diproses : {total_inserted + total_updated}")
print("=======================================================")
for m_name, stat in month_summary.items():
    print(f" - {m_name:10s} : {stat['total']} Penugasan (Baru: {stat['inserted']}, Update: {stat['updated']})")
